import calendar
from datetime import datetime
from openpyxl import load_workbook, Workbook
import requests
from itertools import combinations
from dateutil.relativedelta import relativedelta
from selenium import webdriver
import time
from selenium.webdriver.chrome.options import Options
import json


def main():
    max_comb_number = 5
    residential_token = set_up_residential_token()

    target_wb, filename = filename_input()
    mode_selection = mode_selection_input()
    target_sheet = get_sheet(target_wb)
    row_index = 32
    start_month, start_year, end_month, end_year = get_date_range()

    start_date = datetime(start_year, start_month, 1).date()
    end_date = datetime(end_year, end_month, calendar.monthrange(
        end_year, end_month)[1]).date()

    today = datetime.today().date()
    if start_date <= datetime(today.year-3, today.month, 1).date():
        print('Warning: Midland Residentials only accept search within 3 years. It still proceeds to search other properties within the date range for you!')

    start_date = str(start_date)
    end_date = str(end_date)

    result_wb_path = 'sample_result.xlsx'

    while target_sheet[f'H{row_index}'].value is not None:

        target_prop_name = target_sheet[f'G{row_index}'].value
        area_list = get_area_list(target_sheet, row_index)

        usages = get_usages_list(target_sheet, row_index)
        usage_price_dict = make_usage_price_dict(
            target_sheet, row_index, usages)

        for usage in usages:
            for t in tx_type:
                if all(v is not None for v in usage_price_dict[usage][t].values()):
                    if usage == 'Commercial' or usage == 'Office':
                        prepare_sheet_for_com_and_office(target_wb, target_sheet, filename, row_index, t, usage, area_list,
                                                         start_date, end_date, usage_price_dict, target_prop_name, result_wb_path, mode_selection, max_comb_number)

                    if usage == 'Residential' or usage == 'Apartments':
                        prepare_sheet_for_residential(target_wb, target_sheet, filename, row_index, t, usage, area_list, start_date,
                                                      end_date, usage_price_dict, target_prop_name, result_wb_path, residential_token, mode_selection, max_comb_number)

        for column_name in column_name_list:
            if target_sheet[f'{column_name}{row_index}'].value is None:
                target_sheet[f'{column_name}{row_index}'] = 'NA'
                
        row_index += 1


def filename_input():
    filename = input("What's your file's name?")
    try:
        target_wb = load_workbook(filename)
        return target_wb, filename
    except:
        print("Your file name is wrong! Please enter again!")
        filename_input()


def mode_selection_input():
    mode_selection = input(
        'Please choose a mode to run on (single/ multiple)?')
    if mode_selection != 'single' and mode_selection != 'multiple':
        mode_selection_input()
    return mode_selection


def set_up_residential_token():
    chrome_options = Options()
    chrome_options.add_argument('--headless')
    chrome_options.add_argument('--disable-gpu')

    try:
        driver = webdriver.Chrome('./chromedriver', options=chrome_options)
        driver.get('https://www.midland.com.hk/en/list/transaction')
        time.sleep(1)
        cookies = driver.get_cookies()
        token = next(
            filter(lambda cookie_dict: cookie_dict['name'] == 'token', cookies))['value']
        driver.close()
        return token

    except:
        raise NoDriverError


def get_sheet(wb, error=False, sname=''):
    if error:
        sname = input(
            f'There is no sheet called {sname}, please re-enter the name! (default: C1p.1)')
    else:
        sname = input(
            "What's the name of the sheet you're working at? (default: C1p.1)")
    if sname == '':
        target_sheet = wb['C1p.1']
        return target_sheet
    else:
        try:
            target_sheet = wb[sname]
            return target_sheet
        except:
            get_sheet(wb, error=True, sname=sname)


def get_date_range(error=False):
    if error:
        print('Please input the correct format: "MM/YYYY"!!!!')
    raw_start = input(
        f'Please input the start month and year of the search (format: MM/YYYY)')
    raw_end = input(
        f'Please input the end month and year of the search (format: MM/YYYY)')
    try:
        raw_start_split = raw_start.split('/')
        raw_end_split = raw_end.split('/')
        start_month = int(raw_start_split[0])
        end_month = int(raw_end_split[0])
        start_year = int(raw_start_split[1])
        end_year = int(raw_end_split[1])
        return [start_month, start_year, end_month, end_year]
    except:
        get_date_range(error=True)


def get_area_list(target_sheet, row_index):
    raw_area = target_sheet[f'H{row_index}'].value
    area_list = []
    for a in raw_area.replace(' And ', ' & ').replace(' AND ', ' & ').replace(' and ', ' & ').split('&'):
        area_list.append(a.strip())
    return area_list


def get_usages_list(target_sheet, row_index):
    raw_usage = target_sheet[f'I{row_index}'].value

    usages = []

    for r_usage in raw_usage.split('/'):
        r_usage = r_usage.strip()
        if r_usage not in ['Apartments', 'Residential', 'Office', 'Commercial']:
            print(
                f'Usage in row {row_index} not valid. Only Apartments, Residential, Office and Commercial are available!')
            continue
        if r_usage == 'Apartments':
            r_usage = 'Residential'
        usages.append(r_usage)

    return list(set(usages))


def make_usage_price_dict(target_sheet, row_index, usages):
    usage_price_dict = {}

    for usage in usages:
        price_range_dict = {
            'rental': {
                'upper': None,
                'lower': None,
            },
            'selling': {
                'upper': None,
                'lower': None
            }
        }
        for t in tx_type:
            if t == 'rental':
                raw_rental_valuation_cell = f"{usage_col_dict[usage][t]['valuation']}{row_index}"
                rental_valuation = str(target_sheet[raw_rental_valuation_cell].value) if (
                    target_sheet[raw_rental_valuation_cell].value != 'NA') and (target_sheet[raw_rental_valuation_cell].value is not None) else None

                raw_rental_actual_cell = f'{usage_col_dict[usage][t]["actual"]}{row_index}'
                rental_actual = str(target_sheet[raw_rental_actual_cell].value) if (
                    target_sheet[raw_rental_actual_cell].value != 'NA') and (target_sheet[raw_rental_actual_cell].value is not None) else None
                if (rental_valuation is None) ^ (rental_actual is None):
                    print(
                        f'Actual or valuation price misses in row {row_index} Rental Price Valuation Summary')
                    continue
                if (rental_valuation is not None) and (rental_actual is not None):
                    valuation_range_list = [float(limit.strip().replace(
                        ',', '')) for limit in rental_valuation.split('-')]
                    actual = float(rental_actual.strip().replace(',', ''))

                    valuation_range = [valuation_range_list[0]
                                       * 0.9, valuation_range_list[-1]*1.1]
                    actual_range = [actual * 0.9, actual * 1.1]
                    if actual_range[1] <= valuation_range[0] or valuation_range[1] <= actual_range[0]:
                        price_range_dict[t]['upper'] = actual_range[1]
                        price_range_dict[t]['lower'] = actual_range[0]
                    elif actual_range[0] >= valuation_range[0] and actual_range[1] <= valuation_range[1]:
                        price_range_dict[t]['upper'] = actual_range[1]
                        price_range_dict[t]['lower'] = actual_range[0]
                    elif actual_range[0] <= valuation_range[0] and actual_range[1] >= valuation_range[0] and actual_range[1] <= valuation_range[1]:
                        price_range_dict[t]['upper'] = actual_range[1]
                        price_range_dict[t]['lower'] = valuation_range[0]
                    elif actual_range[0] >= valuation_range[0] and actual_range[0] <= valuation_range[1] and actual_range[1] >= valuation_range[1]:
                        price_range_dict[t]['upper'] = valuation_range[1]
                        price_range_dict[t]['lower'] = actual_range[0]
                    elif actual_range[1] >= valuation_range[1] and valuation_range[0] >= actual_range[0]:
                        price_range_dict[t]['upper'] = actual_range[1]
                        price_range_dict[t]['lower'] = actual_range[0]

            elif t == 'selling':
                raw_selling_cell = f"{usage_col_dict[usage][t]}{row_index}"
                selling_valuation = float(str(target_sheet[raw_selling_cell].value).strip().replace(',', '')) if (
                    target_sheet[raw_selling_cell].value != 'NA') and (target_sheet[raw_selling_cell].value is not None) else None
                if selling_valuation is not None:
                    price_range_dict[t]['upper'] = selling_valuation*1.1
                    price_range_dict[t]['lower'] = selling_valuation*0.9

            usage_price_dict[usage] = price_range_dict

    return usage_price_dict


def request_miland_commercial_data(tx, target_area, start_date, end_date, ics_type, usage_price_dict, usage, t, max_trials, comb_number, original_start_date, original_end_date):
    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.83 Safari/537.36'
    }
    url = f'https://www.midlandici.com.hk/ics/property/transaction/json?tx_type%5B%5D={tx}&districts={target_area}&ics_type={ics_type}&lang=english&date_min={start_date}&date_max={end_date}&page_size=1000'

    res = requests.get(url, headers=headers)
    res_json = res.json().get('transactions')
    final_result = None
    final_comb = None
    combs = combinations(res_json, comb_number)

    for comb in combs:
        comb_sum = 0
        count = 0
        target_range = list(usage_price_dict[usage][t].values())
        for c in comb:
            if t == "selling" and c.get('ft_sell') and c['ft_sell'] > 0 and c['ft_sell'] >= min(target_range)*0.8 and c['ft_sell'] <= max(target_range)*1.2:
                comb_sum += c['ft_sell']
                count += 1
            elif t == 'rental' and c.get('ft_rent') and c['ft_rent'] > 0 and c['ft_rent'] >= min(target_range)*0.8 and c['ft_rent'] <= max(target_range)*1.2:
                comb_sum += c['ft_rent']
                count += 1
        if count == comb_number:
            result = comb_sum/comb_number
        else:
            result = -1

        if result <= max(target_range) and result >= min(target_range):
            final_result = result
            final_comb = comb
            break
    if final_result is None or final_comb is None:
        max_trials -= 1
        if max_trials > 0:
            new_start_date = datetime(int(start_date.split('-')[0]), int(start_date.split(
                '-')[1]), int(start_date.split('-')[2])) + relativedelta(months=-1)
            new_start_date = str(new_start_date.date())
            return request_miland_commercial_data(tx, target_area, new_start_date, end_date, ics_type, usage_price_dict, usage, t, max_trials, comb_number, original_start_date, original_end_date)
        elif max_trials == 0:
            new_end_date = datetime(int(end_date.split('-')[0]), int(end_date.split(
                '-')[1]), int(end_date.split('-')[2])) + relativedelta(months=1)
            new_end_date = str(new_end_date.date())
            return request_miland_commercial_data(tx, target_area, start_date, new_end_date, ics_type, usage_price_dict, usage, t, max_trials, comb_number, original_start_date, original_end_date)
        else:
            return [None, None]
    else:
        if comb_number>3:
            comb_number -= 1
            return request_miland_commercial_data(tx, target_area, original_start_date, original_end_date, ics_type, usage_price_dict, usage, t, max_trials, comb_number, original_start_date, original_end_date)
        return [final_result, final_comb]


def prepare_sheet_for_com_and_office(target_wb, target_sheet, filename, row_index, t, usage, area_list, start_date, end_date, usage_price_dict, target_prop_name, result_wb_path, mode_selection, comb_number):
    tx = 'S' if t == 'selling' else 'L'
    if usage == 'Office':
        ics_type = 'c'
    elif usage == 'Commercial':
        ics_type = 's'

    target_area_list = []
    for a in area_list:
        with open('id_region_lookup_commercial_office.json', 'r') as file:
            id_region_lookup_commercial_office = json.load(file)

        if mode_selection == 'single':
            for region_id, region_list in id_region_lookup_commercial_office.items():
                found_region = False

                for reg in region_list:
                    if a == reg:
                        target_area_list.append(str(region_id))
                        found_region = True
                        break

                if found_region:
                    break
        else:
            for region_id, region_list in id_region_lookup_commercial_office.items():
                for reg in region_list:
                    if a == reg:
                        target_area_list.append(str(region_id))

    target_area_list = list(set(target_area_list))

    target_area = '%2C'.join(target_area_list)

    if len(target_area_list) > 0:
        max_trials = 3
        final_result, final_comb = request_miland_commercial_data(
            tx, target_area, start_date, end_date, ics_type, usage_price_dict, usage, t, max_trials, comb_number, start_date, end_date)
    else:
        final_result, final_comb = None, None

    if final_result is not None and final_comb is not None:
        target_sheet[f'{result_col_dict[usage][t]}{row_index}'] = final_result
        target_wb.save(filename)
        try:
            result_wb = load_workbook(result_wb_path)
        except:
            result_wb = Workbook()
        if 'Result Sheet' not in [ws.title for ws in result_wb.worksheets]:
            result_ws = result_wb.create_sheet(title='Result Sheet')
            result_ws['A1'] = f'{usage} {t.capitalize()} Price for {target_prop_name}'
        else:
            result_ws = result_wb['Result Sheet']
            title_row = result_ws.max_row + 2
            result_ws[f'A{title_row}'] = f'{usage} {t.capitalize()} Price for {target_prop_name}'

        start_row = result_ws.max_row + 1
        result_ws[f'A{start_row}'] = 'Transaction Date'
        result_ws[f'B{start_row}'] = 'Market Info'
        result_ws[f'C{start_row}'] = 'Approx. Area'
        result_ws[f'D{start_row}'] = 'Leased Price'
        result_ws[f'E{start_row}'] = 'Sell Price'

        write_row = start_row
        for each in final_comb:
            write_row += 1
            result_ws[f'A{write_row}'] = each.get('tx_date').split(' ')[0]
            result_ws[f'B{write_row}'] = f"{each.get('dist_name_zh')}, {each.get('chi_name')}, {each.get('streetno')}號, {each.get('floor_zh')}, {each.get('flat')}".strip(
            ).strip(',')
            result_ws[f'C{write_row}'] = float(f"{each.get('area')}")
            result_ws[f'D{write_row}'] = int(
                each.get('ft_rent')) if each.get('ft_rent') != 0 else 'NA'
            result_ws[f'E{write_row}'] = int(
                each.get('ft_sell')) if each.get('ft_sell') != 0 else 'NA'

        write_row += 2
        result_ws[f'A{write_row}'] = '(source: Midland)'

        write_row += 2
        result_ws[f'A{write_row}'] = f'Avg {usage} {t.capitalize()} Price per square feet:'

        write_row += 1
        result_ws[f'A{write_row}'] = final_result

        result_wb.save(result_wb_path)
    else:
        print(
            f'Could not find row {row_index} result within 10% difference of valuation price!')


def request_miland_residential_data(tx, target_area, start_date, end_date, usage_price_dict, usage, t, max_trials, token, comb_number, original_start_date, original_end_date):
    start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
    end_date = datetime.strptime(end_date, "%Y-%m-%d").date()

    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.83 Safari/537.36',
        'authorization': f'Bearer {token}'
    }
    url = f'https://data.midland.com.hk/search/v1/transactions?hash=true&lang=zh-hk&currency=HKD&unit=feet&search_behavior=normal&dist_ids={target_area}&tx_date=3year&page=1&limit=100000&tx_type={tx}'

    res = requests.get(url, headers=headers)
    raw_res_json = res.json().get('result')

    def search_date_range(place_dict):
        raw_tx_date = place_dict.get('tx_date')
        if not raw_tx_date:
            return False
        try:
            tx_date = datetime.strptime(
                raw_tx_date.split('T')[0], "%Y-%m-%d").date()
            if start_date <= tx_date and tx_date <= end_date:
                return True
            return False
        except:
            return False

    def find_result_and_combo():
        res_json = list(filter(search_date_range, raw_res_json))
        final_result = None
        final_comb = None
        combs = combinations(res_json, comb_number)

        for comb in combs:
            comb_sum = 0
            count = 0
            for c in comb:
                if c.get('price') and c.get('net_area') and c['price'] > 0:
                    comb_sum += c['price']/c['net_area']
                    count += 1
            if count == comb_number:
                result = comb_sum/comb_number
            else:
                result = -1

            target_range = list(usage_price_dict[usage][t].values())
            if result <= max(target_range) and result >= min(target_range):
                final_result = result
                final_comb = comb
                break
        return final_result, final_comb

    final_result, final_comb = find_result_and_combo()

    while (final_result is None or final_comb is None) and (max_trials >= 0):
        if max_trials > 0:
            start_date = start_date + relativedelta(months=-3)
        else:
            end_date = end_date + relativedelta(months=1)
        final_result, final_comb = find_result_and_combo()
        max_trials -= 1

    if comb_number > 3:
        comb_number -= 1
        return request_miland_residential_data(tx, target_area, original_start_date, original_end_date, usage_price_dict, usage, t, max_trials, token, comb_number, original_start_date, original_end_date)

    return final_result, final_comb


def prepare_sheet_for_residential(target_wb, target_sheet, filename, row_index, t, usage, area_list, start_date, end_date, usage_price_dict, target_prop_name, result_wb_path, token, mode_selection, comb_number):
    tx = 'S' if t == 'selling' else 'L'

    target_area_list = []
    for a in area_list:
        with open('id_region_lookup_residential.json', 'r') as file:
            id_region_lookup_residential = json.load(file)

        for region_id, region_list in id_region_lookup_residential.items():
            if mode_selection == 'single':
                found_region = False

                for reg in region_list:
                    if a == reg:
                        target_area_list.append(str(region_id))
                        found_region = True
                        break

                if found_region:
                    break
            else:
                for reg in region_list:
                    if a == reg:
                        target_area_list.append(str(region_id))

    target_area_list = list(set(target_area_list))

    target_area = ','.join(target_area_list)

    if len(target_area_list) > 0:
        max_trials = 3
        final_result, final_comb = request_miland_residential_data(
            tx, target_area, start_date, end_date, usage_price_dict, usage, t, max_trials, token, comb_number, start_date, end_date)
    else:
        final_result, final_comb = None, None

    if final_result is not None and final_comb is not None:
        target_sheet[f'{result_col_dict[usage][t]}{row_index}'] = final_result
        target_wb.save(filename)
        try:
            result_wb = load_workbook(result_wb_path)
        except:
            result_wb = Workbook()
        if 'Result Sheet' not in [ws.title for ws in result_wb.worksheets]:
            result_ws = result_wb.create_sheet(title='Result Sheet')
            result_ws['A1'] = f'{usage} {t.capitalize()} Price for {target_prop_name}'
        else:
            result_ws = result_wb['Result Sheet']
            title_row = result_ws.max_row + 2
            result_ws[f'A{title_row}'] = f'{usage} {t.capitalize()} Price for {target_prop_name}'

        start_row = result_ws.max_row + 1
        result_ws[f'A{start_row}'] = 'Transaction Date'
        result_ws[f'B{start_row}'] = 'Market Info'
        result_ws[f'C{start_row}'] = 'Approx. Area'
        result_ws[f'D{start_row}'] = 'Leased Price'
        result_ws[f'E{start_row}'] = 'Sell Price'

        write_row = start_row
        for each in final_comb:
            write_row += 1
            result_ws[f'A{write_row}'] = each.get('tx_date').split('T')[0]
            address_list = []
            if each.get('subregion'):
                address_list.append(each.get('subregion').get('name'))
            if each.get('estate'):
                address_list.append(each.get('estate').get('name'))
            if each.get('building'):
                address_list.append(each.get('building').get('name'))
            if each.get('floor_level'):
                address_list.append(each.get('floor_level').get('name'))
            result_ws[f'B{write_row}'] = ','.join(
                address_list).strip().strip(',')
            result_ws[f'C{write_row}'] = float(f"{each.get('net_area')}")
            result_ws[f'D{write_row}'] = int(
                each.get('price')/each.get('net_area')) if tx == 'L' else 'NA'
            result_ws[f'E{write_row}'] = int(
                each.get('price')/each.get('net_area')) if tx == 'S' else 'NA'

        write_row += 2
        result_ws[f'A{write_row}'] = '(source: Midland)'

        write_row += 2
        result_ws[f'A{write_row}'] = f'Avg {usage} {t.capitalize()} Price per square feet:'

        write_row += 1
        result_ws[f'A{write_row}'] = final_result

        result_wb.save(result_wb_path)
    else:
        print(
            f'Could not find row {row_index} result within 10% difference of valuation price!')

# district_dict = {
#     "Central": 'CEN',
#     "Western District": 'WES',
#     "Admiralty": "ADM",
#     "Sheung Wan": 'SHW',
#     "Wan Chai": 'WAC',
#     "Wan Chai Waterfront": 'WCN',
#     "Causeway Bay": 'CAB',
#     "North Point": "NOP",
#     "Shau Kei Wan": 'SKW',
#     "Chai Wan": 'CHW',
#     "Quarry Bay": "QUB",
#     "Taikoo Shing": 'TAK',
#     "Siu Sai Wan": "SSW",
#     "Wong Chuk Hang":'WCH',
#     "Aberdeen": "ABE",
#     "Kwai Chung": "KWC",
#     "Tsuen Wan": 'TSW',
#     "Tuen Mun": "TUM",
#     "Yuen Long": 'YUL',
#     "Sheung Shui": 'SHS',
#     "Shek Mun": "SHM",
#     "Sha Tin": "SHT",
#     "Mongkok": 'MOK',
#     "Tsim Sha Tsui": "TST",
#     "Tsim Sha Tsui West": "TSI",
#     "Jordan": 'JOR',
#     "Yau Ma Tei": 'YMT',
#     'Prince Edward': 'PRE',
#     'Tai Kok Tsui': 'TKT',
#     'Tsim Sha Tsui East': 'TSE',
#     'Sham Shui Po': 'SSP',
#     'Cheung Sha Wan': 'CSW',
#     'Kowloon City':'KOC',
#     'Hung Hom': 'HUH',
#     'San Po Kong': 'SPK',
#     "Kwun Tong": 'KWT',
#     'Kowloon Bay': 'KOB'
# }

# district_alias_dict = {}
# for dis, sym in district_dict.items():
#     alias_list = dis.split(' ')
#     alias_new_list = []
#     for index, fragment in enumerate(alias_list):
#         if index > 0:
#             alias_new_list.append(fragment.lower())
#         else:
#             alias_new_list.append(fragment)
#         district_alias_dict[''.join(alias_new_list)] = sym


usage_col_dict = {
    'Commercial': {
        'rental': {
            'valuation': 'X',
            'actual': 'AE'
        },
        'selling': 'AL'
    },
    'Office': {
        'rental': {
            'valuation': 'Y',
            'actual': 'AF'
        },
        'selling': 'AM'
    },
    'Residential': {
        'rental': {
            'valuation': 'Z',
            'actual': 'AG'
        },
        'selling': 'AL'
    }
}

result_col_dict = {
    'Commercial': {
        'rental': "S",
        'selling': 'AO'
    },
    'Office': {
        'rental': "T",
        'selling': 'AP'
    },
    'Residential': {
        'rental': "U",
        'selling': 'AQ'
    }
}

tx_type = ['rental', 'selling']

column_name_list = ['S', 'T', 'U', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AE', 'AF',
                    'AG', 'AH', 'AI', 'AJ', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT']


class NoDriverError(Exception):
    def __init__(self):
        print('There is no chromedriver. Please download it first!')


if __name__ == '__main__':
    main()
